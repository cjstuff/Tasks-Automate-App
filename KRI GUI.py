import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from datetime import datetime
from threading import Thread
import zipfile
import openpyxl
from io import BytesIO
import os


class ExcelProcessingApp:
    def __init__(self, master):
        self.master = master
        self.master.title("KRI Count Processing App")

        # Variables for storing file paths and user inputs
        self.zip_path1 = tk.StringVar()
        self.zip_path2 = tk.StringVar()
        self.selected_month = tk.StringVar(value="January")  # Default selected month is January
        self.year = tk.StringVar(value=datetime.now().year)  # Default year is the current year

        # Create and set up GUI elements
        self.create_widgets()

    def create_widgets(self):
        # Entry for zip file 1
        tk.Label(self.master, text="Query1 Zip File:").grid(row=0, column=0, padx=10, pady=10)
        tk.Entry(self.master, textvariable=self.zip_path1, width=50).grid(row=0, column=1, padx=10, pady=10)
        tk.Button(self.master, text="Browse", command=self.browse_zip1).grid(row=0, column=2, padx=10, pady=10)

        # Entry for zip file 2
        tk.Label(self.master, text="Query2 Zip File:").grid(row=1, column=0, padx=10, pady=10)
        tk.Entry(self.master, textvariable=self.zip_path2, width=50).grid(row=1, column=1, padx=10, pady=10)
        tk.Button(self.master, text="Browse", command=self.browse_zip2).grid(row=1, column=2, padx=10, pady=10)

        # Month input
        tk.Label(self.master, text="Month:").grid(row=2, column=0, padx=10, pady=10)
        months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
                  'November', 'December']
        tk.OptionMenu(self.master, self.selected_month, *months).grid(row=2, column=1, padx=10, pady=10)

        # Year input
        tk.Label(self.master, text="Year:").grid(row=3, column=0, padx=10, pady=10)
        tk.Entry(self.master, textvariable=self.year, width=10).grid(row=3, column=1, padx=10, pady=10)

        # Process Button
        tk.Button(self.master, text="Process", command=self.process_data).grid(row=4, column=1, pady=20)

    def browse_zip1(self):
        file_path = filedialog.askopenfilename(filetypes=[("Zip Files", "*.zip")])
        if file_path:
            self.zip_path1.set(file_path)

    def browse_zip2(self):
        file_path = filedialog.askopenfilename(filetypes=[("Zip Files", "*.zip")])
        if file_path:
            self.zip_path2.set(file_path)

    def process_data(self):
        zip_path1 = self.zip_path1.get()
        zip_path2 = self.zip_path2.get()
        selected_month = self.selected_month.get()
        year = self.year.get()

        # Check if file paths are provided
        if not zip_path1 or not zip_path2:
            messagebox.showerror("Error", "Please select both zip files.")
            return

        # Run processing in a separate thread to keep the GUI responsive
        Thread(target=self.process_in_thread, args=(zip_path1, zip_path2, selected_month, year)).start()

    def process_in_thread(self, zip_path1, zip_path2, selected_month, year):
        try:
            # Open a separate window to show progress
            progress_window = tk.Toplevel(self.master)
            progress_window.title("Processing...")
            progress_label = tk.Label(progress_window, text="Processing data. Please wait...")
            progress_label.pack(pady=20)

            # Process data using the provided function
            process_data(zip_path1, zip_path2, selected_month, year)

            # Close the progress window when processing is complete
            progress_window.destroy()

            # Show completion message
            messagebox.showinfo("Success", "Processing completed. Result saved.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")


def process_data(zip_path1, zip_path2, month, year):
    try:
        # Create a new workbook to store the result
        result_wb = openpyxl.Workbook()

        # Process the first zip file (sheets with 'deviceAction' column)
        process_query1(zip_path1, result_wb, month, year)

        # Process the second zip file (sheets with 'Event Time' and 'message' columns)
        process_query2(zip_path2, result_wb, month, year)

        # Save the result workbook
        result_wb.save(f'Result_{month}_{year}.xlsx')

        print("Processing completed. Result saved.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")


def process_query1(zip_path, result_wb, month, year):
    # Open the zip file (sheets with 'deviceAction' column)
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        # Create a new sheet for the query in the result workbook
        result_sheet = result_wb.create_sheet(title='Query1')

        # Write headers to the result sheet
        result_sheet.append(['Date', 'Blocked Count'])

        # Dictionaries to store blocked counts for each date
        blocked_counts = {}

        # Process each Excel sheet in the zip file
        for file_name in zip_ref.namelist():
            if file_name.endswith('.xlsx'):
                date_str = os.path.splitext(os.path.basename(file_name))[0]

                # Parse date string to datetime object
                try:
                    date = datetime.strptime(f"{month} {date_str}, {year}", "%B %d, %Y").date()
                except ValueError:
                    print(f"Invalid date format for file: {file_name}")
                    continue

                # Read the Excel sheet from the zip file
                with zip_ref.open(file_name) as sheet_file:
                    wb = openpyxl.load_workbook(BytesIO(sheet_file.read()))
                    sheet = wb.active

                    # Get the column index of 'deviceAction'
                    device_action_col_index = None
                    for col_index, col in enumerate(sheet.iter_cols(), 1):
                        if col[0].value == 'deviceAction':
                            device_action_col_index = col_index
                            break

                    if device_action_col_index is not None:
                        # Calculate the count of rows containing 'blocked' in the 'deviceAction' column
                        blocked_count = sum(1 for row in sheet.iter_rows(min_row=2, max_col=device_action_col_index) if
                                            row[device_action_col_index - 1].value == 'blocked')
                        blocked_counts[date] = blocked_count

        # Sort dates in ascending order
        sorted_dates = sorted(blocked_counts.keys())

        # Write the result to the new sheet
        for date in sorted_dates:
            result_sheet.append([date.strftime("%Y-%m-%d"), blocked_counts[date]])

        # Add a row for the total sum of blocked counts
        result_sheet.append(['Total', sum(blocked_counts.values())])


def process_query2(zip_path, result_wb, month, year):
    # Open the zip file (sheets with 'Event Time' and 'message' columns)
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        # Create a new sheet for the query in the result workbook
        result_sheet = result_wb.create_sheet(title='Query2')

        # Write headers to the result sheet
        result_sheet.append(['Date', 'Total Event Time Count', 'Filtered Message Count'])

        # Dictionaries to store counts for each date
        total_event_time_counts = {}
        filtered_message_counts = {}

        # Process each Excel sheet in the zip file
        for file_name in zip_ref.namelist():
            if file_name.endswith('.xlsx'):
                date_str = os.path.splitext(os.path.basename(file_name))[0]

                # Parse date string to datetime object
                try:
                    date = datetime.strptime(f"{month} {date_str}, {year}", "%B %d, %Y").date()
                except ValueError:
                    print(f"Invalid date format for file: {file_name}")
                    continue

                # Read the Excel sheet from the zip file
                with zip_ref.open(file_name) as sheet_file:
                    wb = openpyxl.load_workbook(BytesIO(sheet_file.read()))
                    sheet = wb.active

                    # Get the column indices of 'Event Time' and 'message'
                    event_time_col_index = 1  # Assuming 'Event Time' is in column A
                    message_col_index = 12  # Assuming 'message' is in column L

                    # Calculate the total count of 'Event Time'
                    total_event_time_count = sheet.max_row - 1  # Exclude header row
                    total_event_time_counts[date] = total_event_time_count

                    # Calculate the count of rows containing specified words in 'message' column
                    filtered_message_count = sum(
                        1 for row in sheet.iter_rows(min_row=2, max_col=message_col_index)
                        if any(word in str(row[message_col_index - 1].value).lower() for word in
                               ['tcp', 'sql injection', 'brute force'])
                    )
                    filtered_message_counts[date] = filtered_message_count

        # Sort dates in ascending order
        sorted_dates = sorted(total_event_time_counts.keys())

        # Write the result to the new sheet
        for date in sorted_dates:
            result_sheet.append(
                [date.strftime("%Y-%m-%d"), total_event_time_counts.get(date, 0), filtered_message_counts.get(date, 0)])

        # Add a row for the total sum of counts
        result_sheet.append(['Total', sum(total_event_time_counts.values()), sum(filtered_message_counts.values())])


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessingApp(root)
    root.mainloop()
