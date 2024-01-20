import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
import pandas as pd


def process_excel(input_path):
    # Load the matrix sheet from the existing Excel file
    matrix_df = pd.read_excel(input_path, sheet_name='matrix')

    # Task: Check if one user has two different countries
    # and ignore rows where 'Attacker Address' or 'Attacker Geo Country Name' is blank
    # duplicate_users = matrix_df.duplicated(subset=['Attacker User ID', 'Attacker Geo Country Name'], keep=False)
    # users_with_two_countries = matrix_df[duplicate_users & ~matrix_df['Attacker Address'].isnull() &
    #                                      ~matrix_df['Attacker Geo Country Name'].isnull()]['Attacker User ID']
    users_with_two_countries = matrix_df[matrix_df.duplicated(subset=['Attacker User ID'], keep=False) &
                                         ~matrix_df['Attacker Address'].isnull() &
                                         ~matrix_df['Attacker Geo Country Name'].isnull()]['Attacker User ID'].unique()
    # Create a new Excel workbook and add a sheet for users with two different countries
    wb = Workbook()
    ws_users_with_two_countries = wb.active
    ws_users_with_two_countries.title = 'UsersWithTwoCountries'

    # Write headers to the sheet
    headers = ['End Time', 'Attacker User ID', 'Attacker User Name', 'Attacker Address',
               'Attacker Geo Country Name', 'Name', 'Device Action']
    ws_users_with_two_countries.append(headers)

    # Write data for users with two different countries
    prev_user_id = None
    for user_id in users_with_two_countries:
        user_data = matrix_df[(matrix_df['Attacker User ID'] == user_id) &
                              ~matrix_df['Attacker Address'].isnull() &
                              ~matrix_df['Attacker Geo Country Name'].isnull()]
        if len(user_data['Attacker Geo Country Name'].unique()) > 1:
            if prev_user_id is not None:
                ws_users_with_two_countries.append([])  # Add a blank row when 'Attacker User ID' changes
            for _, row in user_data.iterrows():
                ws_users_with_two_countries.append(list(row[['End Time', 'Attacker User ID', 'Attacker User Name',
                                                             'Attacker Address', 'Attacker Geo Country Name',
                                                             'Name', 'Device Action']]))
            prev_user_id = user_id

    # Convert 'Attacker Username' column to numbers
    for row in ws_users_with_two_countries.iter_rows(min_row=2, max_col=2,
                                                     max_row=ws_users_with_two_countries.max_row):
        for cell in row:
            try:
                cell.value = int(cell.value)
            except (ValueError, TypeError):
                pass  # Ignore non-integer values

    # Create a new sheet for unique Attacker User ID
    unique_user_ids = matrix_df[matrix_df['Attacker User ID'].isin(users_with_two_countries)][
        'Attacker User ID'].unique()
    ws_unique_user_ids = wb.create_sheet(title='UniqueAttackerUserIDs')
    ws_unique_user_ids.append(['Attacker User ID'])
    for user_id in unique_user_ids:
        user_data = matrix_df[(matrix_df['Attacker User ID'] == user_id) &
                              ~matrix_df['Attacker Geo Country Name'].isnull()]
        if len(user_data['Attacker Geo Country Name'].unique()) > 1:
            ws_unique_user_ids.append([user_id])

    # Save the workbook with the new sheets
    output_path = 'Matrix_output.xlsx'
    wb.save(output_path)
    return output_path


class ExcelProcessingApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Matrix Processing App")

        # Variables for storing file paths
        self.input_path = tk.StringVar()

        # Create and set up GUI elements
        self.create_widgets()

    def create_widgets(self):
        # Entry for input Excel file
        tk.Label(self.master, text="Input Excel File:").grid(row=0, column=0, padx=10, pady=10)
        tk.Entry(self.master, textvariable=self.input_path, width=50).grid(row=0, column=1, padx=10, pady=10)
        tk.Button(self.master, text="Browse", command=self.browse_input).grid(row=0, column=2, padx=10, pady=10)

        # Process Button
        tk.Button(self.master, text="Process", command=self.process_data).grid(row=1, column=1, pady=20)

    def browse_input(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.input_path.set(file_path)

    def process_data(self):
        input_path = self.input_path.get()

        # Check if the input file path is provided
        if not input_path:
            tk.messagebox.showerror("Error", "Please select the input Excel file.")
            return

        # Process the Excel file
        output_path = process_excel(input_path)

        # Show completion message
        tk.messagebox.showinfo("Success", f"Processing completed. Result saved to {output_path}")


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessingApp(root)
    root.mainloop()
