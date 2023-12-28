import zipfile
import openpyxl
from io import BytesIO
from datetime import datetime
import os

def process_data(zip_path1, zip_path2, month, year):
    try:
        # Create a new workbook to store the result
        result_wb = openpyxl.Workbook()

        # Process the first zip file (sheets with 'deviceAction' column)
        process_query1(zip_path1, result_wb, month, year)

        # Process the second zip file (sheets with 'Event Time' and 'message' columns)
        process_query2(zip_path2, result_wb, month, year)

        # Save the result workbook
        result_wb.save(f'Result_{month}_{year}.xlsx')

        print("Processing completed. Result saved.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

def process_query1(zip_path, result_wb, month, year):
    # Open the zip file (sheets with 'deviceAction' column)
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        # Create a new sheet for the query in the result workbook
        result_sheet = result_wb.create_sheet(title='Query1')

        # Write headers to the result sheet
        result_sheet.append(['Date', 'Blocked Count'])

        # Dictionaries to store blocked counts for each date
        blocked_counts = {}

        # Process each Excel sheet in the zip file
        for file_name in zip_ref.namelist():
            if file_name.endswith('.xlsx'):
                date_str = os.path.splitext(os.path.basename(file_name))[0]

                # Parse date string to datetime object
                try:
                    date = datetime.strptime(f"{month} {date_str}, {year}", "%B %d, %Y").date()
                except ValueError:
                    print(f"Invalid date format for file: {file_name}")
                    continue

                # Read the Excel sheet from the zip file
                with zip_ref.open(file_name) as sheet_file:
                    wb = openpyxl.load_workbook(BytesIO(sheet_file.read()))
                    sheet = wb.active

                    # Get the column index of 'deviceAction'
                    device_action_col_index = None
                    for col_index, col in enumerate(sheet.iter_cols(), 1):
                        if col[0].value == 'deviceAction':
                            device_action_col_index = col_index
                            break

                    if device_action_col_index is not None:
                        # Calculate the count of rows containing 'blocked' in the 'deviceAction' column
                        blocked_count = sum(1 for row in sheet.iter_rows(min_row=2, max_col=device_action_col_index) if row[device_action_col_index - 1].value == 'blocked')
                        blocked_counts[date] = blocked_count

        # Sort dates in ascending order
        sorted_dates = sorted(blocked_counts.keys())

        # Write the result to the new sheet
        for date in sorted_dates:
            result_sheet.append([date.strftime("%Y-%m-%d"), blocked_counts[date]])

        # Add a row for the total sum of blocked counts
        result_sheet.append(['Total', sum(blocked_counts.values())])

def process_query2(zip_path, result_wb, month, year):
    # Open the zip file (sheets with 'Event Time' and 'message' columns)
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        # Create a new sheet for the query in the result workbook
        result_sheet = result_wb.create_sheet(title='Query2')

        # Write headers to the result sheet
        result_sheet.append(['Date', 'Intrusion Count', 'N/w Breach Count'])

        # Dictionaries to store counts for each date
        total_event_time_counts = {}
        filtered_message_counts = {}

        # Process each Excel sheet in the zip file
        for file_name in zip_ref.namelist():
            if file_name.endswith('.xlsx'):
                date_str = os.path.splitext(os.path.basename(file_name))[0]

                # Parse date string to datetime object
                try:
                    date = datetime.strptime(f"{month} {date_str}, {year}", "%B %d, %Y").date()
                except ValueError:
                    print(f"Invalid date format for file: {file_name}")
                    continue

                # Read the Excel sheet from the zip file
                with zip_ref.open(file_name) as sheet_file:
                    wb = openpyxl.load_workbook(BytesIO(sheet_file.read()))
                    sheet = wb.active

                    # Get the column indices of 'Event Time' and 'message'
                    event_time_col_index = 1  # Assuming 'Event Time' is in column A
                    message_col_index = 12  # Assuming 'message' is in column L

                    # Calculate the total count of 'Event Time'
                    total_event_time_count = sheet.max_row - 1  # Exclude header row
                    total_event_time_counts[date] = total_event_time_count

                    # Calculate the count of rows containing specified words in 'message' column
                    filtered_message_count = sum(
                        1 for row in sheet.iter_rows(min_row=2, max_col=message_col_index)
                        if any(word in str(row[message_col_index - 1].value).lower() for word in ['tcp', 'sql injection', 'brute force'])
                    )
                    filtered_message_counts[date] = filtered_message_count

        # Sort dates in ascending order
        sorted_dates = sorted(total_event_time_counts.keys())

        # Write the result to the new sheet
        for date in sorted_dates:
            result_sheet.append([date.strftime("%Y-%m-%d"), total_event_time_counts.get(date, 0), filtered_message_counts.get(date, 0)])

        # Add a row for the total sum of counts
        result_sheet.append(['Total', sum(total_event_time_counts.values()), sum(filtered_message_counts.values())])

# Example usage:
zip_path1 = 'b002.zip'  # Replace with the path to the first zip file
zip_path2 = 'Query2.zip'  # Replace with the path to the second zip file
month = 'November'  # Replace with the actual month
year = '2023'  # Replace with the actual year

process_data(zip_path1, zip_path2, month, year)
